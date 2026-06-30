"""Export highlight/underline entries for Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from pdf_editor.document import TextMarkupEntry

_HEADER_FILL = PatternFill(
    start_color="E2EFDA",
    end_color="E2EFDA",
    fill_type="solid",
)
_CELL_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_CONTENT_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGNMENT = Alignment(vertical="top")

_PAGE_COLUMN_PX = 70
_ANNOTATION_COLUMN_PX = 50
_CONTENT_COLUMN_PX = 800


def _column_width_from_pixels(pixels: int) -> float:
    """Approximate Excel column width (character units) from pixel width."""
    return (pixels - 5) / 7


def _rgb_to_excel_hex(rgb: tuple[float, float, float]) -> str:
    channels = (
        max(0, min(255, int(round(value * 255))))
        for value in rgb[:3]
    )
    return "".join(f"{channel:02X}" for channel in channels)


def _annotation_label(kind: str) -> str:
    return "U" if kind == "underline" else "H"


def export_markup_entries_to_xlsx(entries: list[TextMarkupEntry], path: str | Path) -> None:
    """Write *entries* as an Excel workbook with header row [페이지, 주석, 내용]."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "형광펜 밑줄"

    rows: list[tuple[int, str, str, tuple[float, float, float] | None]] = []
    for entry in entries:
        rows.append(
            (
                entry.page_index + 1,
                _annotation_label(entry.kind),
                entry.text,
                entry.rgb,
            )
        )

    headers = ["페이지", "주석", "내용"]
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.border = _CELL_BORDER
        cell.fill = _HEADER_FILL
        cell.alignment = _TOP_ALIGNMENT

    for row_index, (page, label, text, rgb) in enumerate(rows, start=2):
        page_cell = sheet.cell(row=row_index, column=1, value=page)
        page_cell.border = _CELL_BORDER
        page_cell.alignment = _TOP_ALIGNMENT

        note_cell = sheet.cell(row=row_index, column=2, value=label)
        note_cell.border = _CELL_BORDER
        note_cell.alignment = Alignment(horizontal="center", vertical="center")
        if rgb is not None:
            note_cell.fill = PatternFill(
                start_color=_rgb_to_excel_hex(rgb),
                end_color=_rgb_to_excel_hex(rgb),
                fill_type="solid",
            )

        content_cell = sheet.cell(row=row_index, column=3, value=text)
        content_cell.border = _CELL_BORDER
        content_cell.alignment = _CONTENT_ALIGNMENT

    sheet.column_dimensions["A"].width = _column_width_from_pixels(_PAGE_COLUMN_PX)
    sheet.column_dimensions["B"].width = _column_width_from_pixels(_ANNOTATION_COLUMN_PX)
    sheet.column_dimensions["C"].width = _column_width_from_pixels(_CONTENT_COLUMN_PX)

    workbook.save(Path(path))
